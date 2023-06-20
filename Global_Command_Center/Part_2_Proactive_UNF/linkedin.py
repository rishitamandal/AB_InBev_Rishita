import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import smtplib


# LinkedIn profile URL
linkedin_url = "https://www.linkedin.com/in/user_url"

# Excel file path
excel_file_path = "historical_data.xlsx"

# Gmail SMTP configuration
smtp_server = "smtp.gmail.com"
smtp_port = 587
sender_email = "sender-email"
sender_password = "sener-password"
recipient_email = "receipient-email"


def fetch_linkedin_data():
    # Set up Selenium options
    options = Options()
    options.add_argument("--headless")  # Run in headless mode (without GUI)
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--no-sandbox")

    # Set up the Chrome driver service
    service = Service("C:/Users/Ranvi/Desktop/p2/chromedriver")  # Replace with the actual path to chromedriver executable
    driver = webdriver.Chrome(service=service, options=options)

    # Navigate to LinkedIn profile
    driver.get(linkedin_url)

    # Wait for the page to load (adjust the sleep time if needed)
    time.sleep(5)

    # Find and extract the number of unread messages and notifications (modify the locators as needed)
    unread_messages = driver.find_element(By.XPATH, "your_xpath_for_unread_messages").text
    unread_notifications = driver.find_element(By.XPATH, "your_xpath_for_unread_notifications").text

    # Close the browser
    driver.quit()

    return int(unread_messages), int(unread_notifications)


def get_previous_data():
    # Load the Excel file
    workbook = openpyxl.load_workbook(excel_file_path)

    # Select the appropriate sheet (modify as needed)
    sheet = workbook["Sheet1"]

    # Get the previous occurrence data from the last row of the sheet
    previous_unread_messages = sheet.cell(row=sheet.max_row, column=1).value
    previous_unread_notifications = sheet.cell(row=sheet.max_row, column=2).value

    # Close the Excel file
    workbook.close()

    return previous_unread_messages, previous_unread_notifications


def save_current_data(unread_messages, unread_notifications):
    # Load the Excel file
    workbook = openpyxl.load_workbook(excel_file_path)

    # Select the appropriate sheet (modify as needed)
    sheet = workbook["Sheet1"]

    # Append the current data to a new row
    row = [unread_messages, unread_notifications]
    sheet.append(row)

    # Save the Excel file
    workbook.save(excel_file_path)

    # Close the Excel file
    workbook.close()


def compare_data(current_messages, current_notifications, previous_messages, previous_notifications):
    comparison_result = {
        "message_change": current_messages - previous_messages,
        "notification_change": current_notifications - previous_notifications
    }
    return comparison_result


def send_email(num_messages, num_notifications, comparison_result):
    email_subject = "LinkedIn Notifications Update"

    # Create the email message with HTML body
    message = MIMEMultipart("alternative")
    message["Subject"] = email_subject
    message["From"] = sender_email
    message["To"] = recipient_email

def send_email(num_messages, num_notifications, comparison_result):
    # Generate the HTML body
    email_body = f"""
    <html>
        <body>
        <h2>LinkedIn Notifications Update</h2>
            <p>Number of unread messages: {num_messages}</p>
            <p>Number of unread notifications: {num_notifications}</p>
            <h3>Comparison:</h3>
            <p>Unread messages change: {comparison_result['message_change']}</p>
            <p>Unread notifications change: {comparison_result['notification_change']}</p>
        </body>
    </html>
    """

    return html_body

def send_email(num_messages, num_notifications, comparison_result):
    email_subject = "LinkedIn Notifications Update"

    # Create the email message with HTML body
    message = MIMEMultipart("alternative")
    message["Subject"] = email_subject
    message["From"] = sender_email
    message["To"] = recipient_email

    # Generate the HTML body
    email_body = generate_html_body(num_messages, num_notifications, comparison_result)

    # Attach the HTML body to the email
    message.attach(MIMEText(email_body, "html"))

    # Connect to Gmail SMTP server and send the email
    with smtplib.SMTP(smtp_server, smtp_port) as server:
        server.starttls()
        server.login(sender_email, sender_password)
        server.sendmail(sender_email, recipient_email, message.as_string())

